import os
import win32com.client
import re
import json
import csv
from datetime import datetime
from openai import OpenAI
from typing import Dict, List, Optional, Tuple
import time

class OutlookEmailProcessor:
    def __init__(self, api_key: str, folder_name: str = "Inbox", poll_interval: int = 60):
        self.api_key = api_key
        self.folder_name = folder_name
        self.poll_interval = poll_interval
        self.processed_emails = set()
        self.case_types = self._load_case_types()
        self.client = OpenAI(api_key=api_key)
        
    def _load_case_types(self) -> Dict[str, List[str]]:
        """Load case types from Cases.txt file"""
        case_types = {}
        current_case = None
        
        try:
            cases_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Cases.txt')
            if os.path.exists(cases_file):
                with open(cases_file, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if line.startswith('Case:'):
                            current_case = line[5:].strip()
                            case_types[current_case] = []
                        elif current_case and line and not line.startswith('Case Type:'):
                            # Only add the line if it's not empty and not a Case Type header
                            if line and not line.startswith('Case Type:'):
                                case_types[current_case].append(line.strip())
            
            if not case_types:
                raise FileNotFoundError("No valid case types found in Cases.txt")
                
            print("Loaded case types:")
            for case, subcases in case_types.items():
                print(f"- {case}: {', '.join(subcases) if subcases else 'No subcases'}")
                
                
        except Exception as e:
            print(f"Error loading case types: {str(e)}")
            case_types = {"Other": ["General"]}
            
        return case_types

    def _extract_with_gpt4(self, email_subject: str, email_body: str, email_date=None) -> Dict[str, str]:
        """Extract information from email using GPT-4"""
        try:
            # Format the cases and subcases for the prompt
            case_selection_text = ""
            for case, subcases in self.case_types.items():
                if subcases:
                    subcase_text = "\n    " + "\n    ".join([f"- {subcase}" for subcase in subcases])
                    case_selection_text += f"- {case}:{subcase_text}\n"
                else:
                    case_selection_text += f"- {case} (No subcases)\n"
            
            # Prepare the prompt for GPT-4
            prompt = f"""Extract the following information from this email. If information is not found, use the default values provided:

1. Store Number: A 5-digit number starting with 1, 2, 4, 6, or 8. If not found, use 89999.
2. Contact Person: First name only. If not found, use "Staff".
3. Callback Phone Number: Any phone number in the email. If not found, leave empty.
4. Case: Select the SINGLE MOST APPROPRIATE case from this EXACT list (case-sensitive):
{case_selection_text}
5. Case Type: For the selected case, choose the SINGLE MOST APPROPRIATE subcase from the indented list under that case. If no subcase matches, leave this empty.
6. Summary: A concise summary in 8 words or less.

Email Subject: {email_subject}
Email Body: {email_body[:4000]}  # Limit to first 4000 chars to avoid token limits

Return the response in JSON format with these exact keys: store_number, contact_person, phone_number, case, case_type, summary.

IMPORTANT RULES:
- The 'case' field MUST be EXACTLY one of the main cases shown above (case-sensitive)
- The 'case_type' field MUST be EXACTLY one of the subcases listed under the selected case, or an empty string if none match
- DO NOT make up or modify any case names or subcase names
- If the email doesn't clearly match any case, use the first case as default
- If the email doesn't clearly match any subcase, leave case_type as an empty string"""
            
            response = self.client.chat.completions.create(
                model="gpt-4",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.2,
                max_tokens=500
            )
            
            # Parse the response
            result = response.choices[0].message.content.strip()
            try:
                # Try to parse as JSON
                data = json.loads(result)
                # Ensure all required fields are present
                for key in ['store_number', 'contact_person', 'phone_number', 'case_type', 'date', 'time', 'summary']:
                    if key not in data:
                        data[key] = self._get_default_values()[key]
                return data
            except json.JSONDecodeError:
                # If JSON parsing fails, use the fallback method
                print("Warning: Failed to parse GPT-4 response as JSON, using fallback extraction")
                return self._extract_from_text(result, email_subject, email_body)
                
        except Exception as e:
            print(f"Error processing email with GPT-4: {str(e)}")
            return self._get_default_values()

    def _extract_from_text(self, text: str, subject: str, body: str) -> Dict[str, str]:
        """Fallback method to extract information from text if JSON parsing fails"""
        result = self._get_default_values()
        
        # Extract store number (5 digits starting with 1,2,4,6,8)
        store_match = re.search(r'\b([12468]\d{4})\b', subject + " " + body)
        if store_match:
            result['store_number'] = store_match.group(1)
            
        # Extract contact person (first word after "Contact" or "From" in the body)
        contact_match = re.search(r'(?:Contact|From)[:\s]+(\w+)', body, re.IGNORECASE)
        if contact_match:
            result['contact_person'] = contact_match.group(1)
            
        # Extract phone number (simple pattern)
        phone_match = re.search(r'(\d{3}[-.]?\d{3}[-.]?\d{4})', body)
        if phone_match:
            result['phone_number'] = phone_match.group(1)
            
        # Extract case type (look for case types in the text)
        for case_type in self.case_types:
            if case_type.lower() in body.lower() or case_type.lower() in subject.lower():
                result['case_type'] = case_type
                break
                
        # Use current date and time as fallback
        now = datetime.now()
        result['date'] = now.strftime('%m/%d/%y')
        result['time'] = now.strftime('%I:%M %p').lstrip('0')
        
        # Generate a simple summary
        result['summary'] = f"{subject[:100]}..."
            
        return result

    def _get_default_values(self) -> Dict[str, str]:
        """Return default values for all fields"""
        now = datetime.now()
        # Get the first available case and its first subcase as default, or use 'Other'
        default_case = next(iter(self.case_types.keys()), 'Other')
        default_case_type = self.case_types.get(default_case, [''])[0] if self.case_types.get(default_case) else ''
        
        return {
            'store_number': '89999',
            'contact_person': 'Staff',
            'phone_number': '',
            'case': default_case,
            'case_type': default_case_type,
            'date': now.strftime('%m/%d/%y'),
            'time': now.strftime('%I:%M %p').lstrip('0'),
            'summary': 'Unable to process email content'
        }

    def process_email(self, mail_item) -> Optional[Dict[str, str]]:
        """Process a single email"""
        try:
            email_id = mail_item.EntryID
            if email_id in self.processed_emails:
                return None
                
            print(f"\nProcessing email: {mail_item.Subject}")
            
            # Get email content and metadata
            subject = mail_item.Subject or "No Subject"
            body = mail_item.Body or ""
            
            # Get date and time from email
            if hasattr(mail_item, 'ReceivedTime') and mail_item.ReceivedTime is not None:
                received_time = mail_item.ReceivedTime
                date_str = received_time.strftime('%m/%d/%y')
                # Windows-compatible time formatting
                time_str = received_time.strftime('%I:%M %p')
                if time_str.startswith('0'):
                    time_str = time_str[1:]  # Remove leading zero
            else:
                now = datetime.now()
                date_str = now.strftime('%m/%d/%y')
                time_str = now.strftime('%I:%M %p')
                if time_str.startswith('0'):
                    time_str = time_str[1:]  # Remove leading zero
            
            # Process with GPT-4 (without date/time in the prompt)
            result = self._extract_with_gpt4(subject, body)
            
            if not result:
                return None
                
            # Add email metadata including our formatted date and time
            result['date'] = date_str
            result['time'] = time_str
            result['email_id'] = email_id
            result['subject'] = subject
            
            # Mark as read
            mail_item.UnRead = False
            mail_item.Save()
            
            # Save to CSV
            self.save_to_csv(result)
            
            # Mark as processed
            self.processed_emails.add(email_id)
            
            return result
            
        except Exception as e:
            print(f"Error processing email: {str(e)}")
            return None

    def save_to_csv(self, data: Dict[str, str], output_file: str = r'C:\Users\nnyamekye\CascadeProjects\windsurf-project\email_report.xlsx'):
        """Save extracted data to Excel file with table formatting"""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
            import os

            # Prepare the data
            formatted_data = {
                'Store Number': [data.get('store_number', data.get('Store Number', ''))],
                'Contact Person': [data.get('contact_person', data.get('Contact Person', ''))],
                'Phone Number': [data.get('phone_number', data.get('Phone Number', ''))],
                'Case': [data.get('case', data.get('Case', ''))],
                'Case Type': [data.get('case_type', data.get('Case Type', ''))],
                'Date': [data.get('date', data.get('Date', ''))],
                'Time': [data.get('time', data.get('Time', ''))],
                'Summary': [data.get('summary', data.get('Summary', ''))]
            }

            # Ensure directory exists
            os.makedirs(os.path.dirname(os.path.abspath(output_file)) or '.', exist_ok=True)

            # Create or append to Excel file
            if os.path.exists(output_file):
                # Read existing data
                existing_df = pd.read_excel(output_file, engine='openpyxl')
                # Append new data
                updated_df = pd.concat([existing_df, pd.DataFrame(formatted_data)], ignore_index=True)
            else:
                updated_df = pd.DataFrame(formatted_data)

            # Save to Excel with table formatting
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                updated_df.to_excel(writer, index=False, sheet_name='Email Reports')
                worksheet = writer.sheets['Email Reports']
                
                # Create a table
                from openpyxl.worksheet.table import Table, TableStyleInfo
                table = Table(displayName="EmailTable", ref=f"A1:{get_column_letter(len(updated_df.columns))}{len(updated_df) + 1}")
                
                # Add a default style with striped rows and banded columns
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                     showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                table.tableStyleInfo = style
                worksheet.add_table(table)
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Cap width at 50

            print(f"\nUpdated Excel file: {output_file}")

        except Exception as e:
            print(f"Error saving to Excel file: {str(e)}")
            import traceback
            traceback.print_exc()

    def monitor_emails(self):
        """Monitor the specified Outlook folder for new emails"""
        print(f"Starting to monitor '{self.folder_name}' folder...")
        
        try:
            outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
            
            # Get the target folder
            if self.folder_name.lower() == "inbox":
                folder = outlook.GetDefaultFolder(6)  # 6 is the Inbox
            else:
                # Try to get a subfolder
                inbox = outlook.GetDefaultFolder(6)
                folder = inbox.Folders[self.folder_name]
                
            print(f"Monitoring folder: {folder.Name}")
            
            # Initial load of existing emails (mark as processed without processing)
            messages = folder.Items
            for message in messages:
                self.processed_emails.add(message.EntryID)
            
            print(f"Found {len(self.processed_emails)} existing emails. Monitoring for new ones...")
            
            while True:
                try:
                    # Get all unread emails
                    messages = folder.Items
                    messages.Sort("[ReceivedTime]", True)  # Sort by received time, newest first
                    new_emails = 0
                    
                    # Process unread emails
                    for message in messages:
                        if message.UnRead:
                            result = self.process_email(message)  # save_to_csv is called inside process_email
                            if result:
                                new_emails += 1
                    
                    if new_emails > 0:
                        print(f"Processed {new_emails} new email(s).")
                    
                    print(f"Waiting for new emails (checking every {self.poll_interval} seconds)...")
                    time.sleep(self.poll_interval)
                    
                except KeyboardInterrupt:
                    print("\nStopping email monitor...")
                    break
                except Exception as e:
                    print(f"Error in monitoring loop: {str(e)}")
                    time.sleep(60)  # Wait a minute before retrying
                    
        except Exception as e:
            print(f"Error setting up Outlook: {str(e)}")
            print("Make sure Outlook is installed and running.")
        finally:
            print("Email monitoring stopped.")

def main():
    print("Outlook Email Processor with GPT-4")
    print("--------------------------------")
    
    # Get OpenAI API key
    api_key = "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"
    if not api_key:
        api_key = "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"
    
    if not api_key:
        print("Error: OpenAI API key is required.")
        return
    
    # Get folder name (default: Inbox)
    folder_name = "Generated Responses" or "Inbox"
    
    # Get poll interval
    while True:
        try:
            poll_interval = 15
            if poll_interval < 10:
                print("Poll interval must be at least 10 seconds.")
                continue
            break
        except ValueError:
            print("Please enter a valid number.")
    
    # Create and start the email processor
    try:
        processor = OutlookEmailProcessor(
            api_key=api_key,
            folder_name=folder_name,
            poll_interval=poll_interval
        )
        
        print("\nStarting email monitoring. Press Ctrl+C to stop.")
        processor.monitor_emails()
        
    except Exception as e:
        print(f"Error: {str(e)}")
        print("Make sure all dependencies are installed and Outlook is running.")
    
    print("Program ended.")

if __name__ == "__main__":
    main()
